from statistics import mean

import numpy as np
import openpyxl
import statistics

from noncenteral_tstatic_cdf import bad_data, outlier_test, kk, tolerance, reasonable


def eliminate_bad_data(x, x0, x1):
    # Stage 1 -- Bad data elimination
    y = bad_data(x, x0, x1)
    if y is None:
        return None, "All intervals eliminated at bad data stage"
    return y, None

def eliminate_outliers(y):
    # Stage 2 -- Outlier elimination
    z = outlier_test(y)
    if z is None:
        return None, "All intervals eliminated at outlier elimination stage"
    return z, None

def calculate_mean_std(z):
    # Calculate mean and standard deviation
    mean_left = np.mean(z[:, 0])
    std_left = statistics.stdev(z[:, 0])
    mean_right = np.mean(z[:, 1])
    std_right = statistics.stdev(z[:, 1])
    return mean_left, std_left, mean_right, std_right


def tolerance_limit_processing(z, mean_left, std_left, mean_right, std_right, a, significance_level):
    # Stage 3 -- Tolerance limit processing
    tolerance_factor = kk(a, significance_level, len(z) + 1)
    y0 = tolerance(z[:, 0], mean_left, std_left, tolerance_factor)
    y1 = tolerance(z[:, 1], mean_right, std_right, tolerance_factor)
    z = []

    for i in range(len(y0)):
        if int(y0[i]) != -1000 and int(y1[i]) != -1000:
            z.append([y0[i], y1[i]])

    if z:
        return np.array(z), None
    else:
        return np.empty((0, 2)), "All intervals eliminated at tolerance limit processing stage"


def reasonable_interval_processing(z, mean_left, std_left, mean_right, std_right):
    # Stage 4 -- Reasonable interval processing
    z = reasonable(z.tolist(), mean_left, mean_right, std_left, std_right)
    if z:
        return np.array(z), None
    else:
        return np.empty((0, 2)), "All intervals eliminated at reasonable interval processing stage"



def dataclean(x, x0, x1, a, significance_level):
    # Preprocess raw interval data for a given word to eliminate unacceptable intervals using all of the above tests
    y = bad_data(x, x0, x1)
    if y is None:
        return "All intervals eliminated at bad data stage"

    z, message = eliminate_outliers(y)
    if message:
        return message

    mean_left, std_left, mean_right, std_right = calculate_mean_std(z)

    z, message = tolerance_limit_processing(z, mean_left, std_left, mean_right, std_right, a, significance_level)
    if message:
        return message

    z, message = reasonable_interval_processing(z, mean_left, std_left, mean_right, std_right)
    if message:
        return message

    out = [[], [], []]
    out[0] = z
    # Compute sample means of residual interval endpoints
    out[1] = np.mean(z[:, 0])
    out[2] = np.mean(z[:, 1])
    return out





# ws = openpyxl.load_workbook('excel/depreciated/Very Bad interval data.xlsx')["Sheet1"]
# x_very_bad = []
# for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=2, values_only=True):
#     x_very_bad.append([float(cell) for cell in row if cell is not None])
#
# ws = openpyxl.load_workbook('excel/depreciated/Bad interval data.xlsx')['Sheet1']
# x_bad = []
# for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=2, values_only=True):
#     x_bad.append([float(cell) for cell in row if cell is not None])
#
# ws = openpyxl.load_workbook('excel/depreciated/Somewhat Bad interval data.xlsx')['Sheet1']
# x_somewhat_bad = []
# for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=2, values_only=True):
#     x_somewhat_bad.append([float(cell) for cell in row if cell is not None])
#
# ws = openpyxl.load_workbook('excel/depreciated/Fair interval data.xlsx')['Sheet1']
# x_fair = []
# for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=2, values_only=True):
#     x_fair.append([float(cell) for cell in row if cell is not None])
#
# ws = openpyxl.load_workbook('excel/depreciated/Somewhat Good interval data.xlsx')['Sheet1']
# x_somewhat_good = []
# for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=2, values_only=True):
#     x_somewhat_good.append([float(cell) for cell in row if cell is not None])
#
# ws = openpyxl.load_workbook('excel/depreciated/Good interval data.xlsx')['Sheet1']
# x_good = []
# for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=2, values_only=True):
#     x_good.append([float(cell) for cell in row if cell is not None])
#
# ws = openpyxl.load_workbook('excel/depreciated/Very Good interval data.xlsx')['Sheet1']
# x_very_good = []
# for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=2, values_only=True):
#     x_very_good.append([float(cell) for cell in row if cell is not None])
#
#
#
# yVB = dataclean(x_very_bad, 0, 10, 0.05, 0.05)[0]
# yB = dataclean(x_bad, 0, 10, 0.05, 0.05)[0]
# ySB = dataclean(x_somewhat_bad, 0, 10, 0.05, 0.05)[0]
# yF = dataclean(x_fair, 0, 10, 0.05, 0.05)[0]
# ySG = dataclean(x_somewhat_good, 0, 10, 0.05, 0.05)[0]
# yG = dataclean(x_good, 0, 10, 0.05, 0.05)[0]
# yVG = dataclean(x_very_good, 0, 10, 0.05, 0.05)[0]